#!/usr/bin/env python3
"""
Convert Astrum_Full_Pricelist_*.xlsx to a single CSV for Urban Focus import.

Usage:
  python scripts/convert_astrum_pricelist.py "C:/path/to/Astrum_Full_Pricelist_SRP_Plus20_Rounded.xlsx"
  python scripts/convert_astrum_pricelist.py   # defaults to storage/imports source

Output: storage/imports/astrum_pricelist.csv
Columns: sku,name,model_number,description,images,price,srp_price,stock,category,warranty
  - price = storefront price from the workbook (taken as-is)
  - srp_price = original Astrum SRP (reference only)
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: python -m pip install openpyxl")
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = ROOT / "storage" / "imports" / "astrum_pricelist.csv"

SHEET_CATEGORY = {
    "Smart": "Gadgets",
    "Connect": "USB Peripherals",
    "Listen": "Speakers",
    "PowerUp": "Power Banks",
    "Replace": "Laptop Screens",
}

SKIP_SHEETS = {"Menu"}


def col_index(header_row: tuple, label: str) -> int | None:
    label = label.lower().strip()
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        if str(cell).strip().lower() == label:
            return i
    return None


def parse_replace_sheet(ws, category: str, writer, stats: dict) -> None:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "part no" in cells and "price" in cells:
            header_idx = i
            break

    if header_idx is None:
        return

    header = rows[header_idx]
    idx_sku = col_index(header, "part no")
    idx_name = col_index(header, "compatible part no") or col_index(header, "part no")
    idx_price = col_index(header, "price")
    idx_srp = col_index(header, "srp")
    idx_size = col_index(header, "size")

    if idx_sku is None or idx_price is None:
        return

    for row in rows[header_idx + 1 :]:
        if not row or idx_sku >= len(row):
            continue
        sku = str(row[idx_sku] or "").strip()
        if not sku or sku.lower() in {"part no", "laptop replacement batteries internal (12 months warranty)"}:
            continue

        price = row[idx_price] if idx_price < len(row) else None
        if price is None or str(price).strip() == "":
            continue

        try:
            price_f = float(price)
        except (TypeError, ValueError):
            continue
        if price_f <= 0:
            continue

        srp = row[idx_srp] if idx_srp is not None and idx_srp < len(row) else None
        try:
            srp_f = float(srp) if srp not in (None, "") else ""
        except (TypeError, ValueError):
            srp_f = ""

        size = str(row[idx_size] or "").strip() if idx_size is not None and idx_size < len(row) else ""
        name = f'{size} Replacement Screen'.strip() if size else sku
        if idx_name is not None and idx_name < len(row) and row[idx_name]:
            extra = str(row[idx_name]).strip()
            if extra and extra != sku:
                name = f"{name} — {extra[:120]}"

        writer.writerow(
            {
                "sku": sku,
                "name": name[:250],
                "model_number": "",
                "description": name[:500],
                "images": "",
                "price": f"{price_f:.2f}",
                "srp_price": f"{srp_f:.2f}" if srp_f != "" else "",
                "stock": "0",
                "category": category,
                "warranty": "12 Months",
            }
        )
        stats["written"] += 1


def parse_standard_sheet(ws, category: str, writer, stats: dict) -> None:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if not row:
            continue
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if cells[:4] == ["part no", "model no", "image", "name"] or (
            "part no" in cells and "price" in cells and "name" in cells
        ):
            header_idx = i
            break

    if header_idx is None:
        return

    header = rows[header_idx]
    idx_sku = col_index(header, "part no")
    idx_model = col_index(header, "model no")
    idx_image = col_index(header, "image")
    idx_name = col_index(header, "name")
    idx_desc = col_index(header, "description")
    idx_opt = col_index(header, "options")
    idx_price = col_index(header, "price")
    idx_srp = col_index(header, "srp")
    idx_warranty = col_index(header, "warranty")

    if idx_sku is None or idx_price is None:
        return

    carry_name = ""
    carry_desc = ""
    carry_model = ""
    carry_category = category

    for row in rows[header_idx + 1 :]:
        if not row:
            continue

        def cell(i: int | None):
            if i is None or i >= len(row):
                return None
            return row[i]

        sku_raw = cell(idx_sku)
        if sku_raw is None or str(sku_raw).strip() == "":
            continue

        sku = str(sku_raw).strip()
        price = cell(idx_price)
        if price is None or str(price).strip() == "":
            if sku and " " in sku and not any(ch.isdigit() for ch in sku[:6]):
                carry_category = sku
            continue

        try:
            price_f = float(price)
        except (TypeError, ValueError):
            continue
        if price_f <= 0:
            continue

        name = str(cell(idx_name) or "").strip()
        desc = str(cell(idx_desc) or "").strip()
        model = str(cell(idx_model) or "").strip()
        opt = str(cell(idx_opt) or "").strip() if idx_opt is not None else ""

        if name:
            carry_name = name
            carry_desc = desc or carry_desc
            carry_model = model or carry_model
        else:
            name = carry_name
            desc = carry_desc or desc
            model = carry_model or model

        if not name:
            name = sku

        if opt:
            name = f"{name} ({opt})"

        image = str(cell(idx_image) or "").strip()
        if image and not image.startswith("http"):
            image = ""

        srp = cell(idx_srp)
        try:
            srp_f = float(srp) if srp not in (None, "") else ""
        except (TypeError, ValueError):
            srp_f = ""

        warranty = str(cell(idx_warranty) or "").strip() or "12 Months"

        writer.writerow(
            {
                "sku": sku,
                "name": name[:250],
                "model_number": model[:80],
                "description": (desc or name)[:2000],
                "images": image,
                "price": f"{price_f:.2f}",
                "srp_price": f"{srp_f:.2f}" if srp_f != "" else "",
                "stock": "0",
                "category": carry_category,
                "warranty": warranty,
            }
        )
        stats["written"] += 1


def convert(path: Path, out: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    out.parent.mkdir(parents=True, exist_ok=True)
    stats = {"written": 0, "sheets": 0}

    fieldnames = [
        "sku",
        "name",
        "model_number",
        "description",
        "images",
        "price",
        "srp_price",
        "stock",
        "category",
        "warranty",
    ]

    with out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            category = SHEET_CATEGORY.get(sheet_name, sheet_name)
            stats["sheets"] += 1

            if sheet_name == "Replace":
                parse_replace_sheet(ws, category, writer, stats)
            else:
                parse_standard_sheet(ws, category, writer, stats)

    wb.close()
    return stats


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if src is None:
        downloads = Path.home() / "Downloads"
        matches = sorted(downloads.glob("Astrum_Full_Pricelist*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not matches:
            print("Pass path to Astrum xlsx file.")
            raise SystemExit(1)
        src = matches[0]

    if not src.is_file():
        print(f"File not found: {src}")
        raise SystemExit(1)

    stats = convert(src, DEFAULT_OUT)
    print(f"Converted {src.name}")
    print(f"  Sheets: {stats['sheets']}")
    print(f"  Products: {stats['written']}")
    print(f"  Output: {DEFAULT_OUT}")


if __name__ == "__main__":
    main()
